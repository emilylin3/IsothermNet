import os
import sys
import shutil
import subprocess
import math
import pandas as pd
import numpy as np
from pymatgen.io.cif import CifFile,CifParser
import xlsxwriter as xl
import openpyxl

path_job = "/home/elin49/RASPA/simulations/bin/"
path_qmof_cif = "/home/elin49/RASPA/simulations/share/raspa/structures/cif/"          # Path for all the relaxed qmof .cif files    

# Remove non-.___ files - output folder
def delNonFile(fileType,dirPath):
    outputFiles = os.listdir(dirPath)
    indx_vec = []
    for indx in range(len(outputFiles)):
        if fileType not in outputFiles[indx]:
            indx_vec.append(indx)
    for i in sorted(indx_vec,reverse=True):
        del outputFiles[i]
    return outputFiles

# Automate SLURM scheduler
def automateSlurm(ext_P,qmof_cif,HeVF):
    simulation_fileName = "simulation.input"
    moleculeName = "CO2"

    # Framework path (.cif)
    frameworkName = qmof_cif.replace(".cif","")                                            # Same name as structure (.cif file) - change accordingly

    # Calculate unit cells
    VDW_cutoff = 12              # Van der Waal cutoff limit [Angstrom]
    [key,value] = list(CifParser(path_qmof_cif+qmof_cif).as_dict().items())[0]
    unit_cell_a = float(value.get("_cell_length_a"))                         # Unit cell length of MOF (x) [Angstrom] - obtained from .cif file
    unit_cell_b = float(value.get("_cell_length_b"))                         # Unit cell length of MOF (y) [Angstrom] - obtained from .cif file
    unit_cell_c = float(value.get("_cell_length_c"))                         # Unit cell length of MOF (z) [Angstrom] - obtained from .cif file

    angle_a = float(value.get("_cell_angle_alpha"))                         # Angle of MOF (a) [deg]
    angle_b = float(value.get("_cell_angle_beta"))                         # Angle of MOF (b) [deg]
    angle_c = float(value.get("_cell_angle_gamma"))                         # Angle of MOF (c) [deg]

    tempd=(math.cos(angle_a*np.pi/180)-math.cos(angle_c*np.pi/180)*math.cos(angle_b*np.pi/180))/math.sin(angle_c*np.pi/180)
    ax=unit_cell_a; bx=unit_cell_b*math.cos(angle_c*np.pi/180); cx=unit_cell_c*math.cos(angle_b*np.pi/180)
    ay=0; by=unit_cell_b*math.sin(angle_c*np.pi/180); cy=unit_cell_c*tempd
    az=0; bz=0; cz=unit_cell_c*math.sqrt(1-((math.cos(angle_b*np.pi/180))**2)-(tempd**2))

    # Calculate vector products of cell vectors
    axb1=(ay*bz)-(az*by); axb2=(az*bx)-(ax*bz); axb3=(ax*by)-(ay*bx)
    bxc1=by*cz-bz*cy; bxc2=bz*cx-bx*cz; bxc3=bx*cy-by*cx
    cxa1=cy*az-ay*cz; cxa2=ax*cz-az*cx; cxa3=ay*cx-ax*cy

    # Calculate volume of cell
    vol_cell=abs(ax*bxc1+ay*bxc2+az*bxc3)

    # Calculate cell perpendicular widths
    pwa=vol_cell/math.sqrt(bxc1*bxc1+bxc2*bxc2+bxc3*bxc3)
    pwb=vol_cell/math.sqrt(cxa1*cxa1+cxa2*cxa2+cxa3*cxa3)
    pwc=vol_cell/math.sqrt(axb1*axb1+axb2*axb2+axb3*axb3)

    a = math.ceil(VDW_cutoff*2/pwa)
    b = math.ceil(VDW_cutoff*2/pwb)
    c = math.ceil(VDW_cutoff*2/pwc)

    n_cycles = 6000            
    n_initCycles = 4000        
    printStep = 500
    
    bool_contAftCrash = "no"
    n_binarySteps = 1000
    bool_restart = "no"
    
    ff_pseudo_dir = "MOF_ff_pseudo_v2"     
    bool_useCharge = "yes"

    ext_T = 298                                        # External temperature [K]

    moleculeDef_dir = "MoleculeDefinitions"
    fugCoeff = 1.0
    transProb = 0.5
    rotProb = 0.5
    reinsProb = 0.5
    swapProb = 1.0
    createNoMol = 0

    with open(path_job+simulation_fileName,'w') as f:
        f.write("SimulationType                MonteCarlo"); f.write("\n")
        f.write("NumberOfCycles                {}".format(n_cycles)); f.write("\n")
        f.write("NumberOfInitializationCycles  {}".format(n_initCycles)); f.write("\n")
        f.write("PrintEvery                    {}".format(printStep)); f.write("\n"); f.write("\n")

        f.write("RestartFile                   {}".format(bool_restart)); f.write("\n"); f.write("\n")
                
        f.write("Forcefield                    {}".format(ff_pseudo_dir)); f.write("\n")
        f.write("UseChargesFromCIFFile         {}".format(bool_useCharge)); f.write("\n"); f.write("\n")

        f.write("Framework 0"); f.write("\n")
        f.write("FrameworkName {}".format(frameworkName)); f.write("\n")
        f.write("UnitCells {} {} {}".format(a,b,c)); f.write("\n")
        f.write("HeliumVoidFraction {}".format(HeVF)); f.write("\n")
        f.write("ExternalTemperature {}".format(ext_T)); f.write("\n")
        f.write("ExternalPressure {}".format(ext_P)); f.write("\n"); f.write("\n")

        f.write("Component 0 MoleculeName             {}".format(moleculeName)); f.write("\n")
        f.write("            MoleculeDefinition       {}".format(moleculeDef_dir)); f.write("\n")
        f.write("            FugacityCoefficient      {}".format(fugCoeff)); f.write("\n")
        f.write("            TranslationProbability   {}".format(transProb)); f.write("\n")
        f.write("            RotationProbability      {}".format(rotProb)); f.write("\n")
        f.write("            ReinsertionProbability   {}".format(reinsProb)); f.write("\n")
        f.write("            SwapProbability          {}".format(swapProb)); f.write("\n")
        f.write("            CreateNumberOfMolecules  {}".format(createNoMol))

    os.chdir(path_job)              # location of .job file
    subprocess.run(["./simulate"])
    
    return moleculeName, ext_T, a, b, c

def sciNot(pressure):
    P_str = str("{:e}".format(pressure))
    indx1 = P_str.find('.')
    indx2 = P_str.find('e')
    
    n = 0
    for i in range(indx1,indx2):
        if P_str[i] == str(0):
            n += 1
    
    if n == indx2-indx1-1:
        P_new = P_str.replace(P_str[indx2-1:indx1+(indx2-indx1-n)-1:-1],"")
        P_new = P_new.replace(P_str[P_str.find('.')],"")
    elif n < indx2-indx1-1:
        P_new = P_str.replace(P_str[indx2-1:indx1+(indx2-indx1-n)-1:-1],"") 
        
    return P_new

# MAIN CODE: calculate isotherm for qmofs already w/ calculated He void fraction 
outputFileDir = "/home/elin49/RASPA/simulations/bin/Output/System_0/"  

ext_T = 298  
pressure_ref = [1e3,5e3,1e4,5e4,1e5,2e5,3e5,4e5,5e5,7e5,1e6,1.5e6,2e6,2.5e6,3e6,3.5e6,4e6,4.5e6,5e6]
pressure = [7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5,7e5]

indx = sys.argv[1]
P = pressure[int(indx)-1]
moleculeName = "CO2"  

outputFiles = delNonFile(".data",outputFileDir)

row = 0

indx_ref = np.where(np.array(pressure_ref)==P)[0][0]

data_track = pd.read_excel(path_job+"trackProgress_CO2_"+str(int(indx_ref+1))+".xlsx")
data = pd.read_excel(path_job+"HeVF_CO2.xlsx")
for i in range(len(data_track.iloc[:,0])):
    if str(data_track.iloc[:,0][i]) != "nan":
        qmof = data_track.iloc[:,0][i]+".cif"
        frameworkName = data_track.iloc[:,0][i]

        VDW_cutoff = 12              # Van der Waal cutoff limit [Angstrom]
        [key,value] = list(CifParser(path_qmof_cif+qmof).as_dict().items())[0]
        unit_cell_a = float(value.get("_cell_length_a"))                         # Unit cell length of MOF (x) [Angstrom] - obtained from .cif file
        unit_cell_b = float(value.get("_cell_length_b"))                         # Unit cell length of MOF (y) [Angstrom] - obtained from .cif file
        unit_cell_c = float(value.get("_cell_length_c"))                         # Unit cell length of MOF (z) [Angstrom] - obtained from .cif file

        angle_a = float(value.get("_cell_angle_alpha"))                         # Angle of MOF (a) [deg]
        angle_b = float(value.get("_cell_angle_beta"))                         # Angle of MOF (b) [deg]
        angle_c = float(value.get("_cell_angle_gamma"))                         # Angle of MOF (c) [deg]

        tempd=(math.cos(angle_a*np.pi/180)-math.cos(angle_c*np.pi/180)*math.cos(angle_b*np.pi/180))/math.sin(angle_c*np.pi/180)
        ax=unit_cell_a; bx=unit_cell_b*math.cos(angle_c*np.pi/180); cx=unit_cell_c*math.cos(angle_b*np.pi/180)
        ay=0; by=unit_cell_b*math.sin(angle_c*np.pi/180); cy=unit_cell_c*tempd
        az=0; bz=0; cz=unit_cell_c*math.sqrt(1-((math.cos(angle_b*np.pi/180))**2)-(tempd**2))

        # Calculate vector products of cell vectors
        axb1=(ay*bz)-(az*by); axb2=(az*bx)-(ax*bz); axb3=(ax*by)-(ay*bx)
        bxc1=by*cz-bz*cy; bxc2=bz*cx-bx*cz; bxc3=bx*cy-by*cx
        cxa1=cy*az-ay*cz; cxa2=ax*cz-az*cx; cxa3=ay*cx-ax*cy

        # Calculate volume of cell
        vol_cell=abs(ax*bxc1+ay*bxc2+az*bxc3)

        # Calculate cell perpendicular widths
        pwa=vol_cell/math.sqrt(bxc1*bxc1+bxc2*bxc2+bxc3*bxc3)
        pwb=vol_cell/math.sqrt(cxa1*cxa1+cxa2*cxa2+cxa3*cxa3)
        pwc=vol_cell/math.sqrt(axb1*axb1+axb2*axb2+axb3*axb3)

        a = math.ceil(VDW_cutoff*2/pwa)
        b = math.ceil(VDW_cutoff*2/pwb)
        c = math.ceil(VDW_cutoff*2/pwc)

        if P < 1e6:
            str_filename = "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data"
        else:
            str_filename = "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data"

        folderName = "isotherm_"+frameworkName+"_"+moleculeName+"_"+str(ext_T)+"K"
        if str_filename not in os.listdir(outputFileDir):
            try: 
                if str_filename not in os.listdir(outputFileDir+folderName):
                    for j in range(len(data.iloc[:,0])):
                        if data_track.iloc[:,0][i] == data.iloc[:,0][j]:
                            HeVF = data.iloc[:,1][j]

                    moleculeName, ext_T, a, b, c = automateSlurm(P,qmof,HeVF)
                    print(str(int(P))+" "+qmof +" | ",flush=True)

                    # Sort files into correct folders (or create folder) - remove HeVF file from correct folder after moving
                    wb = openpyxl.load_workbook(path_job+"trackProgress_CO2_"+str(int(indx_ref+1))+".xlsx")
                    ws = wb["Sheet1"]
                    print("opened workbook",flush=True)

                    folderName = "isotherm_"+frameworkName+"_"+moleculeName+"_"+str(ext_T)+"K"
                    destDir = outputFileDir+folderName

                    if folderName in os.listdir(outputFileDir):
                        if "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data" in os.listdir(outputFileDir):
                            scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data"
                        elif "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data" in os.listdir(outputFileDir):
                            scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data"
                        else:
                            print("Invalid filename")

                    else:
                        os.mkdir(destDir)
                        if "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data" in os.listdir(outputFileDir):
                            scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data"
                        elif "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data" in os.listdir(outputFileDir):
                            scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data"
                        else:
                            print("Invalid filename")

                    shutil.move(scrDir,destDir)
                    ws.cell(row=int(2+i),column=1).value = "nan"
                    wb.save(path_job+"trackProgress_CO2_"+str(int(indx_ref+1))+".xlsx")
                    
            except:
                for j in range(len(data.iloc[:,0])):
                    if data_track.iloc[:,0][i] == data.iloc[:,0][j]:
                        HeVF = data.iloc[:,1][j]

                moleculeName, ext_T, a, b, c = automateSlurm(P,qmof,HeVF)
                print(str(int(P))+" "+qmof +" | ",flush=True)

                # Sort files into correct folders (or create folder) - remove HeVF file from correct folder after moving
                wb = openpyxl.load_workbook(path_job+"trackProgress_CO2_"+str(int(indx_ref+1))+".xlsx")
                ws = wb["Sheet1"]

                folderName = "isotherm_"+frameworkName+"_"+moleculeName+"_"+str(ext_T)+"K"
                destDir = outputFileDir+folderName

                if folderName in os.listdir(outputFileDir):
                    if "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data" in os.listdir(outputFileDir):
                        scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data"
                    elif "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data" in os.listdir(outputFileDir):
                        scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data"
                    else:
                        print("Invalid filename")
                else:
                    os.mkdir(destDir)
                    if "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data" in os.listdir(outputFileDir):
                        scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+str(int(P))+".data"
                    elif "output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data" in os.listdir(outputFileDir):
                        scrDir = outputFileDir+"output_"+frameworkName+"_"+str(a)+"."+str(b)+"."+str(c)+"_"+str(ext_T)+".000000_"+sciNot(P)+".data"
                    else:
                        print("Invalid filename")

                shutil.move(scrDir,destDir)
                ws.cell(row=int(2+i),column=1).value = "nan"
                wb.save(path_job+"trackProgress_CO2_"+str(int(indx_ref+1))+".xlsx")
